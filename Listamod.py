import xlsxwriter
import os
import re  # regular expressions
import datetime
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

# **************************
# ****
def OrdenaVisitas(aVisitas):
    if len(aVisitas) > 3:
        presort = aVisitas[4:]
        return sorted(presort, key=lambda x: x[0], reverse=True)
    return None


# **************************
# ****
def Carrega_PlanilhaMestre(pLista, pWsheet):
    Qtd = 0
    for cLinha in range(3, pWsheet.max_row):
        DataProcessamento = datetime.now
        DataObito = str(pWsheet.cell(cLinha, 2).value)
        NomePaciente = str(pWsheet.cell(cLinha, 3).value)
        Matricula = str(pWsheet.cell(cLinha, 4).value)
        if NomePaciente == "None":
            continue
        pLista.append([NomePaciente, DataObito, Matricula])
        Qtd += 1
    return Qtd


# **************************
# ****
def Filtra_Arquivos(Lista, NomePlanilha="Sheet1"):
    c_Arquivos = 0
    Diretorio = os.listdir()
    for Arquivo in Diretorio:
        if not Arquivo.endswith("xlsx") and not re.search(
            "([12]\d{3}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01]))", Arquivo
        ):
            continue
        # [12]?\d\d\d[12]\d[123]\d
        # ([12]\d{3}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01]))
        try:
            Entrada = load_workbook(Arquivo, data_only=True)
            Folha = Entrada[NomePlanilha]
            Entrada.close()
        except:
            continue

        c_Arquivos += 1
        Lista.append(Arquivo)
    return c_Arquivos


def ProcessaUmDia():
    for linha_dia in range(1, faixa_dia.max_row):
        PacienteDia = str(faixa_dia.cell(linha_dia, 2).value)
        DataVisita = faixa_dia.cell(linha_dia, 1).value
        AvaliacaoDia = faixa_dia.cell(linha_dia, 15).value
        # iLista = iter(Lista)
        ponto = next((x for x in Lista if x[0] == PacienteDia), "None")
        if ponto != "None":

            p = len(ponto)
            if p <= 3:
                ponto.append(1)
                p += 1
            else:
                ponto[3] += 1

            if p >= 4:
                aaaVisita = list((str(DataVisita), AvaliacaoDia))
                ponto.append(aaaVisita)


#                    ponto.append( str( DataVisita ) )
#                    ponto.append( str( AvaliacaoDia ) )
# print( Lista )


ListaDeArquivos = []
if Filtra_Arquivos(ListaDeArquivos) == 0:
    print("* N�o foram encontradas planilhas.\n* Processamento encerrado")
    exit


wb = load_workbook("Planilha Plano B atualizada 30_4_19.xlsx")
ListaPacientes = []
if Carrega_PlanilhaMestre(ListaPacientes, wb["Pacientes"]) == 0:
    print("Banco de pacientes nao encontrado.....")
    exit

for itemArquivo in ListaDeArquivos:
    try:
        itemArquivoDia = load_workbook(itemArquivo, data_only=True)
        itemArquivoFolha = itemArquivoDia["Sheet1"]
    except:
        print("Não consegui abrir " + itemArquivo)
        continue

    for linhaDia in range(1, itemArquivoFolha.max_row):
        PacienteDia = str(itemArquivoFolha.cell(linhaDia, 2).value)
        DataVisita = str(itemArquivoFolha.cell(linhaDia, 1).value)
        AvaliacaoDia = itemArquivoFolha.cell(linhaDia, 3).value
        ponto = next((x for x in ListaPacientes if x[0] == PacienteDia), "None")
        if ponto != "None":
            p = len(ponto)
            if p <= 3:
                ponto.append(1)
                p += 1
            else:
                ponto[3] += 1

            if p >= 4:
                itemVisita = list((str(DataVisita), AvaliacaoDia))
                ponto.append(itemVisita)
"""
for itemPaciente in ListaPacientes:
    ordenado = OrdenaVisitas( itemPaciente ) 
    itemPaciente.pop()
    itemPaciente.append(ordenado)    
"""
print(ListaPacientes)

ub = xlsxwriter.Workbook("wbTestemod.xlsx")
folha = ub.add_worksheet("resultado")

formato_data = ub.add_format({"num_format": "dd/mm/yyyy", "align": "right"})
lin = col = 0

for Linha in ListaPacientes:
    folha.write(lin, 1, Linha[0])
    if Linha[1] != "None":
        date_time = datetime.strptime(Linha[1], "%Y-%m-%d %H:%M:%S")
        folha.write_datetime(lin, 2, date_time, formato_data)

    folha.write(lin, 3, Linha[2])
    for ide in range(3, len(Linha)):
        if Linha[ide] != "None":
            try:
                date_time = datetime.strptime(Linha[id], "%Y-%m-%d %H:%M:%S")
                folha.write_datetime(lin, id + 1, date_time, formato_data)
                folha.write(lin, id + 2, Linha[id + 1])
                id += 2
            except:
                continue
    lin += 1

ub.close()
